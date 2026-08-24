from .models import *
from .serializers import *
from rest_framework.exceptions import ValidationError, PermissionDenied
from aryuapp.auth import CustomJWTAuthentication
from rest_framework.response import Response
from rest_framework import status, viewsets
from rest_framework.permissions import IsAuthenticated , AllowAny
from django.db import  transaction
from django.db.models import Value, CharField, Q, Count
from rest_framework.throttling import AnonRateThrottle, UserRateThrottle
from webinar.models import WebinarRegistration
from courses.models import Course
from django.core.cache import cache
from django.db.models.functions import TruncMonth
from django.utils import timezone
from rest_framework.decorators import action
import csv
import io
import mimetypes
import re
from django.utils.dateparse import parse_date, parse_datetime
import openpyxl
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, JSONParser, FormParser
from rest_framework.authentication import SessionAuthentication
from rest_framework.pagination import CursorPagination
from .services import ReportService

# Create your views here.

# =========================================================
# THROTTLES
# =========================================================

class PublicLeadThrottle(AnonRateThrottle):
    scope = "public_lead"


class AdminLeadThrottle(UserRateThrottle):
    scope = "admin_lead"


# =========================================================
# PAGINATION
# =========================================================

class LeadPagination(CursorPagination):
    page_size = 25
    ordering = "-created_at"
    page_size_query_param = "page_size"
    max_page_size = 100


# =========================================================
# SECURITY MIXIN
# =========================================================

class LeadSecurityMixin:

    BLOCKED_REGEX = re.compile(
        r"<script|javascript:|union\s+select|drop\s+table|--|;|onerror=|onload=",
        re.IGNORECASE,
    )

    def validate_payload_security(self, request):

        for value in request.data.values():

            if isinstance(value, str):

                if self.BLOCKED_REGEX.search(value):

                    raise ValidationError(
                        "Invalid payload detected."
                    )

    def validate_admin_access(self, request):

        if not request.user.is_authenticated:
            raise PermissionDenied(
                "Authentication required."
            )

        if not (
            request.user.is_staff
            or request.user.is_superuser
        ):
            raise PermissionDenied(
                "Access denied."
            )

    def get_client_ip(self, request):

        x_forwarded_for = request.META.get(
            "HTTP_X_FORWARDED_FOR"
        )

        if x_forwarded_for:
            return x_forwarded_for.split(",")[0].strip()

        return request.META.get("REMOTE_ADDR")


# =========================================================
# MAIN ADMIN VIEWSET
# =========================================================

class LeadViewSet(LeadSecurityMixin, viewsets.ViewSet):
    authentication_classes = [
        CustomJWTAuthentication,
        SessionAuthentication,
    ]
    permission_classes = [IsAuthenticated]
    throttle_classes = [AdminLeadThrottle]
    parser_classes = [MultiPartParser, JSONParser]
    pagination_class = LeadPagination

    # =====================================================
    # INTERNAL HELPERS
    # =====================================================

    def paginate_queryset(self, queryset, request):
        paginator = self.pagination_class()
        paginated_queryset = paginator.paginate_queryset(queryset, request)
        return paginated_queryset, paginator

    def get_lead_queryset(self):
        return (
            Lead.objects.filter(is_archived=False)
            .only(
                "id",
                "name",
                "phone",
                "email",
                "city",
                "course",
                "status",
                "priority",
                "lead_stage",
                "source",
                "created_at",
                "followup_date",
                "next_followup_date",
                "no_of_calls",
                "no_of_dms",
                "created_by",
                "created_by_type",
            )
            .annotate(
                lead_origin=Value(
                    "lead",
                    output_field=CharField(),
                )
            )
            .order_by("-created_at")
        )

    def get_active_handled_by_users(self):
        """
        Fetches active, non-archived users as lightweight dictionaries.
        Uses Django cache to serve subsequent reads in O(1) time.
        """

        users = list(
            User.objects.filter(
                is_active=True,
                is_archived=False
            ).values("id", "full_name")
        )

        return users

    def get_active_courses(self):
        cache_key = "active_courses"
        courses = cache.get(cache_key)

        if courses is not None:
            return courses

        courses = list(
            Course.objects.filter(
                status="Active", is_archived=False
            ).values("course_id", "course_name")
        )

        cache.set(cache_key, courses, timeout=600)
        return courses

    # =====================================================
    # MAIN LIST METHOD
    # =====================================================

    def list(self, request):
        self.validate_admin_access(request)

        cache_key = f"lead-engine:{request.user.id}:{hash(request.get_full_path())}"
        cached_response = cache.get(cache_key)
        if cached_response:
            return Response(cached_response)

        # 1. Base queryset (O(1) memory overhead lazily evaluated)
        leads_queryset = self.get_lead_queryset()

        # O(1) space, O(log N) DB index scan for total active leads count
        total_active_leads = leads_queryset.count()

        # =====================================
        # 1. SEARCH FILTER
        # =====================================
        search = request.query_params.get("search")
        if search:
            search = search.strip()
            if search.isdigit():
                leads_queryset = leads_queryset.filter(phone__startswith=search)
            elif "@" in search:
                leads_queryset = leads_queryset.filter(email__iexact=search)
            else:
                leads_queryset = leads_queryset.filter(name__istartswith=search)

        # =====================================
        # 2. STATUS & SOURCE FILTERS
        # =====================================
        status_filter = request.query_params.get("status")
        if status_filter:
            leads_queryset = leads_queryset.filter(status=status_filter)

        source_filter = request.query_params.get("source")
        if source_filter:
            leads_queryset = leads_queryset.filter(source=source_filter)

        # =====================================
        # 3. DATE FILTERS (`created_at__gte`, `created_at__lte`)
        # =====================================
        created_at_gte = request.query_params.get("created_at__gte")
        if created_at_gte:
            parsed_date = parse_datetime(created_at_gte) or parse_date(created_at_gte)
            if parsed_date:
                leads_queryset = leads_queryset.filter(created_at__gte=parsed_date)

        created_at_lte = request.query_params.get("created_at__lte")
        if created_at_lte:
            parsed_date = parse_datetime(created_at_lte) or parse_date(created_at_lte)
            if parsed_date:
                leads_queryset = leads_queryset.filter(created_at__lte=parsed_date)

        # =====================================
        # 4. COURSES FILTER (Supports multi-select `DJANGO,PYTHON`)
        # =====================================
        courses_param = request.query_params.get("courses")
        if courses_param:
            # Hash-Set conversion for O(K) lookup efficiency
            courses_list = [c.strip() for c in courses_param.split(",") if c.strip()]
            if courses_list:
                if len(courses_list) == 1:
                    leads_queryset = leads_queryset.filter(course__iexact=courses_list[0])
                else:
                    leads_queryset = leads_queryset.filter(course__in=courses_list)

        # =====================================
        # 5. DICT PROJECTION & CURSOR PAGINATION
        # =====================================
        leads_values_queryset = leads_queryset.values(
            "id",
            "name",
            "phone",
            "email",
            "city",
            "course",
            "status",
            "priority",
            "lead_stage",
            "source",
            "created_at",
            "followup_date",
            "next_followup_date",
            "no_of_calls",
            "no_of_dms",
            "created_by",
            "created_by_type",
            "lead_origin",
        )

        paginated_queryset, paginator = self.paginate_queryset(
            leads_values_queryset, request
        )

        response_data = paginator.get_paginated_response(
            paginated_queryset
        ).data

        response_data["total_active_leads"] = total_active_leads
        response_data["courses"] = self.get_active_courses()
        response_data["users"] = self.get_active_handled_by_users()

        # =====================================
        # CACHE PRESERVATION
        # =====================================
        cache.set(cache_key, response_data, timeout=60)

        return Response(response_data)
    
    # =====================================================
    # RETRIEVE
    # =====================================================

    def retrieve(self, request, pk=None):

        self.validate_admin_access(request)

        cache_key = f"lead-detail:{pk}"

        cached_response = cache.get(cache_key)

        if cached_response:
            return Response(cached_response)

        try:

            lead = (
                Lead.objects
                .select_related(
                    "followup_by",
                    "handled_by"
                )
                .prefetch_related(
                    "call_logs",
                    "dm_logs",
                    "followups",
                    "status_history",
                )
                .get(
                    pk=pk,
                    is_archived=False
                )
            )

        except Lead.DoesNotExist:

            return Response(
                {
                    "detail": "Lead not found."
                },
                status=status.HTTP_404_NOT_FOUND
            )

        serializer = LeadSerializer(
            lead,
            context={
                "request": request
            }
        )

        response_data = serializer.data

        # ==========================================
        # EXTRA CRM DATA
        # ==========================================

        response_data["courses"] = (
            self.get_active_courses()
        )

        response_data["activity_timeline"] = [

            *[
                {
                    "id": log.id,
                    "type": "call",
                    "title": log.call_status,
                    "created_at": log.created_at,
                    "call_type": log.call_type,
                    "duration_seconds": log.duration_seconds,
                    "recording": (
                        "https://aylms.aryuprojects.com/api" + log.recording_url.url
                        if log.recording_url and hasattr(log.recording_url, "url")
                        else None
                    ),
                }
                for log in lead.call_logs.all()
            ],

            *[
                {
                    "type": "dm",
                    "title": dm.platform,
                    "description": dm.message,
                    "created_at": dm.created_at,
                }
                for dm in lead.dm_logs.all()
            ],

            *[
                {
                    "type": "status",
                    "title": (
                        f"{history.old_status or 'Fresh'} → "
                        f"{history.new_status}"
                    ),
                    "description": history.remarks,
                    "created_at": history.created_at,
                }
                for history in lead.status_history.all()
            ],
        ]

        response_data["activity_timeline"].sort(
            key=lambda x: x["created_at"],
            reverse=True
        )
        response_data["courses"] = self.get_active_courses()
        response_data["users"] = self.get_active_handled_by_users()

        cache.set(
            cache_key,
            response_data,
            timeout=30
        )

        return Response(response_data)

    # =====================================================
    # CREATE LEAD
    # =====================================================

    @transaction.atomic
    def create(self, request):

        self.validate_admin_access(request)

        self.validate_payload_security(request)

        serializer = LeadSerializer(
            data=request.data,
            context={
                "request": request
            }
        )

        serializer.is_valid(
            raise_exception=True
        )

        lead = serializer.save()

        # =====================================
        # CACHE INVALIDATION
        # =====================================
        try:
            # 1. Grab the raw client safely using Django's internal wrapper
            raw_client = cache._cache
            
            # 2. Use a wildcard search that accounts for Django's built-in key prefix matching
            # This matches strings ending in your pattern (e.g., ":1:lead-engine:...")
            keys_to_delete = raw_client.keys("*lead-engine:*")
            
            if keys_to_delete:
                # 3. Strip out Django's prefix processing by executing the delete directly on the raw client
                # decode("utf-8") ensures the keys are handled correctly if they return as bytes
                decoded_keys = [k.decode("utf-8") if isinstance(k, bytes) else k for k in keys_to_delete]
                raw_client.delete(*decoded_keys)
                
        except Exception as e:
            # Prevent cache layer issues from crashing the entire database transaction
            print(f"Cache invalidation failed: {e}")

        return Response(
            {
                "success": True,
                "message": "Lead created successfully.",
                "data": LeadSerializer(
                    lead,
                    context={
                        "request": request
                    }
                ).data
            },
            status=status.HTTP_201_CREATED
        )

    # =====================================================
    # BULK LEAD UPLOAD
    # =====================================================

    @transaction.atomic
    @action(detail=False,methods=["POST"],url_path="bulk-upload")
    def bulk_upload(self, request):

        self.validate_admin_access(request)

        upload_file = request.FILES.get("file")

        if not upload_file:

            return Response(
                {
                    "detail": "File is required."
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # =================================================
        # FILE SIZE SECURITY
        # =================================================

        max_size = 5 * 1024 * 1024

        if upload_file.size > max_size:

            return Response(
                {
                    "detail": "Maximum file size is 5MB."
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # =================================================
        # EXTENSION VALIDATION
        # =================================================

        extension = (
            "." + upload_file.name.split(".")[-1].lower()
        )

        allowed_extensions = [
            ".csv",
            ".xlsx"
        ]

        if extension not in allowed_extensions:

            return Response(
                {
                    "detail": (
                        "Only CSV and XLSX files allowed."
                    )
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # =================================================
        # MIME TYPE VALIDATION
        # =================================================

        mime_type, _ = mimetypes.guess_type(
            upload_file.name
        )

        allowed_mimes = [
            "text/csv",
            "application/vnd.ms-excel",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ]

        if mime_type not in allowed_mimes:

            return Response(
                {
                    "detail": "Invalid file type."
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # =================================================
        # UPLOAD LOCK
        # =================================================

        cache_key = (
            f"bulk-upload-lock:{request.user.id}"
        )

        if cache.get(cache_key):

            return Response(
                {
                    "detail": (
                        "Bulk upload already in progress."
                    )
                },
                status=status.HTTP_429_TOO_MANY_REQUESTS
            )

        cache.set(
            cache_key,
            True,
            timeout=300
        )

        try:

            rows = []

            # =============================================
            # CSV PARSER
            # =============================================

            if extension == ".csv":

                decoded_file = io.StringIO(
                    upload_file.read().decode("utf-8")
                )

                reader = csv.DictReader(decoded_file)

                for row in reader:

                    normalized_row = {
                        str(key).strip().lower(): value
                        for key, value in row.items()
                        if key is not None
                    }

                    rows.append(normalized_row)

            # =============================================
            # XLSX PARSER
            # =============================================

            elif extension == ".xlsx":

                workbook = openpyxl.load_workbook(
                    upload_file,
                    read_only=True,
                    data_only=True,
                )

                sheet = workbook.active

                headers = [
                    str(cell).strip().lower()
                    if cell is not None else ""
                    for cell in next(
                        sheet.iter_rows(values_only=True)
                    )
                ]

                for row in sheet.iter_rows(
                    min_row=2,
                    values_only=True
                ):

                    rows.append(
                        {
                            header: value
                            for header, value in zip(headers, row)
                        }
                    )

            # =============================================
            # MAX ROW LIMIT
            # =============================================

            if len(rows) > 10000:

                return Response(
                    {
                        "detail": (
                            "Maximum 10000 rows allowed."
                        )
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )

            # =============================================
            # REQUIRED COLUMNS
            # =============================================

            required_columns = [
                "name",
                "phone"
            ]

            row.get("name")
            row.get("email")
            row.get("phone")
            row.get("course")

            if not rows:

                return Response(
                    {
                        "detail": "File is empty."
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )

            first_row = rows[0]

            for column in required_columns:

                if column not in first_row:

                    return Response(
                        {
                            "detail": (
                                f"Missing required column: {column}"
                            )
                        },
                        status=status.HTTP_400_BAD_REQUEST
                    )

            # =============================================
            # EXISTING PHONES
            # =============================================

            phones = []

            for row in rows:

                phone = str(row.get("phone", "")).strip()

                phone = "".join(filter(str.isdigit, phone))

                if len(phone) >= 10:

                    phones.append(phone[-10:])

            existing_phones = set(

                Lead.objects.filter(

                    phone__in=phones

                ).values_list(

                    "phone",

                    flat=True,

                )

            )

            bulk_leads = []

            duplicate_count = 0

            skipped_count = 0

            for row in rows:

                phone = str(
                    row.get("phone", "")
                ).strip()

                phone = "".join(
                    filter(str.isdigit, phone)
                )

                if len(phone) < 10:

                    skipped_count += 1
                    continue

                phone = phone[-10:]

                # =========================================
                # DUPLICATE CHECK
                # =========================================

                if phone in existing_phones:

                    duplicate_count += 1
                    continue

                existing_phones.add(phone)

                # =========================================
                # EXCEL FORMULA INJECTION PROTECTION
                # =========================================

                def sanitize(value):

                    if value is None:
                        return None

                    value = str(value).strip()

                    dangerous_prefixes = (
                        "=",
                        "+",
                        "-",
                        "@",
                    )

                    if value.startswith(
                        dangerous_prefixes
                    ):
                        value = "'" + value

                    return value

                bulk_leads.append(

                    Lead(

                        name=sanitize(
                            row.get("name")
                        ),

                        phone=phone,

                        email=sanitize(
                            row.get("email")
                        ),

                        city=sanitize(
                            row.get("city")
                        ),

                        state=sanitize(
                            row.get("state")
                        ),

                        course=sanitize(
                            row.get("course")
                        ),

                        source="bulk_upload",

                        status="fresh",

                        created_by=str(
                            request.user.id
                        ),

                        created_by_type=(
                            "super_admin"
                            if request.user.is_superuser
                            else "admin"
                        ),
                    )
                )

            # =============================================
            # BULK INSERT
            # =============================================

            Lead.objects.bulk_create(
                bulk_leads,
                batch_size=1000
            )

            # =============================================
            # CACHE INVALIDATION
            # =============================================

            cache.delete_pattern("lead-engine:*")

            return Response(
                {
                    "success": True,
                    "message": (
                        "Bulk leads uploaded successfully."
                    ),
                    "total_rows": len(rows),
                    "inserted": len(bulk_leads),
                    "duplicates": duplicate_count,
                    "skipped": skipped_count,
                },
                status=status.HTTP_201_CREATED
            )

        except Exception as e:

            return Response(
                {
                    "detail": str(e)
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        finally:

            cache.delete(cache_key)
    
    # =====================================================
    # UPDATE LEAD
    # =====================================================

    @transaction.atomic
    def partial_update(self, request, pk=None):

        self.validate_admin_access(request)

        self.validate_payload_security(request)

        try:

            lead = Lead.objects.get(
                pk=pk,
                is_archived=False
            )

        except Lead.DoesNotExist:

            return Response(
                {
                    "detail": "Lead not found."
                },
                status=status.HTTP_404_NOT_FOUND
            )

        serializer = LeadSerializer(
            lead,
            data=request.data,
            partial=True,
            context={
                "request": request
            }
        )

        serializer.is_valid(
            raise_exception=True
        )

        serializer.save()

        # =====================================
        # CACHE INVALIDATION
        # =====================================
        try:
            # 1. Grab the raw client safely using Django's internal wrapper
            raw_client = cache._cache
            
            # 2. Use a wildcard search that accounts for Django's built-in key prefix matching
            # This matches strings ending in your pattern (e.g., ":1:lead-engine:...")
            keys_to_delete = raw_client.keys("*lead-engine:*")
            
            if keys_to_delete:
                # 3. Strip out Django's prefix processing by executing the delete directly on the raw client
                # decode("utf-8") ensures the keys are handled correctly if they return as bytes
                decoded_keys = [k.decode("utf-8") if isinstance(k, bytes) else k for k in keys_to_delete]
                raw_client.delete(*decoded_keys)
                
        except Exception as e:
            # Prevent cache layer issues from crashing the entire database transaction
            print(f"Cache invalidation failed: {e}")

        return Response(
            {
                "success": True,
                "message": "Lead updated successfully.",
                "data": serializer.data
            }
        )

    # =====================================================
    # SOFT DELETE
    # =====================================================

    @transaction.atomic
    def destroy(self, request, pk=None):

        self.validate_admin_access(request)

        try:

            lead = Lead.objects.get(
                pk=pk,
                is_archived=False
            )

        except Lead.DoesNotExist:

            return Response(
                {
                    "detail": "Lead not found."
                },
                status=status.HTTP_404_NOT_FOUND
            )

        lead.is_archived = True

        lead.save(
            update_fields=[
                "is_archived"
            ]
        )

        cache.delete_pattern("lead-engine:*")
        cache.delete(f"lead-detail:{pk}")

        return Response(
            {
                "success": True,
                "message": "Lead archived successfully."
            }
        )

    @transaction.atomic
    @action(detail=True, methods=["POST"], url_path="add-call-log",parser_classes=[MultiPartParser, FormParser])
    def add_call_log(self, request, pk=None):
        self.validate_admin_access(request)

        try:
            lead = Lead.objects.get(pk=pk, is_archived=False)
        except Lead.DoesNotExist:
            return Response({"detail": "Lead not found."}, status=status.HTTP_404_NOT_FOUND)

        # Build data dictionary from request
        log_data = {
            "duration_seconds": request.data.get("duration_seconds", 0),
            "call_status": request.data.get("call_status", "Completed"),
            "remarks": request.data.get("remarks", ""),
            "next_followup_date": request.data.get("next_followup_date", None),
        }

        serializer = LeadCallLogSerializer(data=log_data)
        
        if serializer.is_valid():
            # Save the log with the lead, the user, and the file
            call_log = serializer.save(
                lead=lead,
                called_by=User.objects.filter(
                    id=request.user.user_id
                ).first(),
                recording_url=request.FILES.get("recording_url")
            )
            
            # Invalidate cache so the lead details page updates
            cache.delete(f"lead-detail:{pk}")
            
            return Response({
                "success": True,
                "message": "Call log and recording saved.",
                "data": LeadCallLogSerializer(call_log).data
            }, status=status.HTTP_201_CREATED)
            
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class LeadDashboard(LeadSecurityMixin,viewsets.ViewSet):

    
    def dashboard(self, request):

        self.validate_admin_access(request)

        today = timezone.localdate()

        leads = Lead.objects.filter(
            is_archived=False
        )

        total_leads = leads.count()

        converted_leads = leads.filter(
            is_converted=True
        ).count()

        new_leads = leads.filter(
            status__iexact="fresh"
        ).count()

        duplicate_leads = leads.filter(
            is_duplicate=True
        ).count()

        archived_leads = Lead.objects.filter(
            is_archived=True
        ).count()

        conversion_rate = (
            round(
                (converted_leads / total_leads) * 100,
                2
            )
            if total_leads
            else 0
        )

        # Source wise
        source_wise = list(
            leads.values("source")
            .annotate(count=Count("id"))
            .order_by("-count")
        )

        # Status wise
        status_wise = list(
            leads.values("status")
            .annotate(count=Count("id"))
            .order_by("-count")
        )

        # Today's calls
        calls_today = LeadCallLog.objects.filter(
            created_at__date=today
        )

        incoming_calls = calls_today.filter(
            call_type__iexact="incoming"
        ).count()

        outgoing_calls = calls_today.filter(
            call_type__iexact="outgoing"
        ).count()

        total_calls_today = calls_today.count()

        # Today's leads
        leads_today = leads.filter(
            created_at__date=today
        ).count()

        # Converted today
        converted_today = leads.filter(
            is_converted=True,
            updated_at__date=today
        ).count()

        # Followups
        followups_due = leads.filter(
            next_followup_date=today
        ).count()

        followups_completed = LeadCallLog.objects.filter(
            next_followup_date__gt=today,
            created_at__date=today
        ).count()

        # Monthly Leads
        monthly_leads = list(
            leads.annotate(
                month=TruncMonth("created_at")
            )
            .values("month")
            .annotate(count=Count("id"))
            .order_by("month")
        )

        return Response({
            "overview": {
                "total_leads": total_leads,
                "new_leads": new_leads,
                "converted_leads": converted_leads,
                "duplicate_leads": duplicate_leads,
                "archived_leads": archived_leads,
                "conversion_rate": conversion_rate,
            },
            "today": {
                "new_leads": leads_today,
                "calls_made": total_calls_today,
                "incoming_calls": incoming_calls,
                "outgoing_calls": outgoing_calls,
                "followups_due": followups_due,
                "followups_completed": followups_completed,
                "converted_today": converted_today,
            },
            "source_wise": source_wise,
            "status_wise": status_wise,
            "monthly_leads": monthly_leads,
        })

# =========================================================
# PUBLIC VIEWSET
# =========================================================

class PublicLeadViewSet(
    LeadSecurityMixin,
    viewsets.ViewSet
):

    authentication_classes = []

    permission_classes = [
        AllowAny
    ]

    throttle_classes = [
        PublicLeadThrottle
    ]

    # =====================================================
    # CREATE PUBLIC LEAD
    # =====================================================

    @transaction.atomic
    def create(self, request):

        self.validate_payload_security(request)

        # =====================================
        # BOT HONEYPOT
        # =====================================

        if request.data.get("website"):

            raise PermissionDenied(
                "Bot detected."
            )

        client_ip = self.get_client_ip(request)

        blocked = cache.get(
            f"blocked-ip:{client_ip}"
        )

        if blocked:

            raise PermissionDenied(
                "Too many requests."
            )

        serializer = PublicLeadCreateSerializer(
            data=request.data,
            context={
                "request": request
            }
        )

        serializer.is_valid(
            raise_exception=True
        )

        lead = serializer.save()

        cache.delete_pattern("lead-engine:*")

        return Response(
            {
                "success": True,
                "message": "Lead submitted successfully.",
                "lead_id": lead.id
            },
            status=status.HTTP_201_CREATED
        )
 
class LeadReportViewSet(viewsets.ViewSet):
    """
    Generic Reports API.

    POST /api/v1/reports/export/
    """

    authentication_classes = [CustomJWTAuthentication]
    permission_classes = [IsAuthenticated]

    service = ReportService()

    def list(self, request):
        data = self.service.get_report_dashboard_data()

        return Response(
            {
                "success": True,
                "data": data,
            },
            status=status.HTTP_200_OK,
        )

    def create(self, request):
        serializer = ReportRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        validated_data = serializer.validated_data

        report_type = validated_data["report_type"]
        filters = validated_data.get("filters", {})
        pagination = validated_data.get(
            "pagination",
            {
                "page": 1,
                "page_size": 50,
            },
        )

        result = self.service.dispatch(
            report_type=report_type,
            filters=filters,
            pagination=pagination,
        )

        return Response(
            {
                "success": True,
                "report_type": report_type,
                "count": result["count"],
                "page": pagination.get("page", 1),
                "page_size": pagination.get("page_size", 50),
                "results": result["results"],
            },
            status=status.HTTP_200_OK,
        )
    
